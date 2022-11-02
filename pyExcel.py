"""
基于openpyxl的excel读写模块
提供更简单的读写方式
"""
from openpyxl import load_workbook, Workbook


class ExcelWork:

    def __init__(self, filePath):
        """
        初始化, 加载excel，默认选择第一个工作表
        :param filePath: str: 文件地址
        """
        self.filePath = filePath
        try:
            # 加载excel
            self.excel = load_workbook(self.filePath)
        except FileNotFoundError:
            # 创建excel
            self.excel = Workbook(self.filePath)
            # 创建sheet
            self.createSheet('Sheet1')
            # 保存excel
            self.close()
            # 加载excel
            self.excel = load_workbook(self.filePath)
        # sheet设置为第一个工作表
        self.sheet = self.excel.active

    def close(self):
        """
        保存并退出
        :return:
        """
        self.excel.save(self.filePath)

    def createSheet(self, sheetName):
        """
        创建工作表
        :param sheetName: str: 工作表名
        :return:
        """
        self.excel.create_sheet(sheetName)

    def getSheetTitle(self):
        """
        获取当前工作表名称
        :return:  str: 工作表名称
        """
        return self.sheet.title

    def getSheetTitles(self):
        """
        获取excel所有工作表的名称
        :return:  list: [工作表名称,]
        """
        return self.excel.sheetnames

    def delSheet(self):
        """
        删除当前工作表
        :return:
        """
        self.excel.remove(self.sheet)

    def selectSheet(self, sheetName):
        """
        选择工作表, 如果没有将创建
        :param sheetName: str: 工作表名
        :return:
        """
        if sheetName in self.getSheetTitles():
            # 选择工作表
            self.sheet = self.excel[sheetName]
        else:
            # 创建工作表
            self.createSheet(sheetName)
            # 选择工作表
            self.sheet = self.excel[sheetName]

    def setCell(self, r, c, var):
        """
        修改指定行， 列的单元格内容
        :param r: int: 行数
        :param c: int: 列数
        :param var: str: 修改内容
        :return:
        """
        self.sheet.cell(row=r, column=c, value=var)

    def getCell(self, r, c):
        """
        获取指定行， 列的单元格内容
        :param r: int: 行数
        :param c: int: 列数
        :return: str: 单元格内容
        """
        return self.sheet.cell(row=r, column=c).value

    def getRow(self, r):
        """
        获取指定行所有数据
        :param r: int: 行数
        :return: list: [数据,]
        """
        rowList = []
        for cell in self.sheet[r]:
            rowList.append(cell.value)
        return rowList

    def getColumn(self, c):
        """
        获取指定列所有数据
        :param c: int: 列数
        :return: list: [数据,]
        """
        columnList = []
        for temp in range(1, self.sheet.max_row + 1):
            columnList.append(self.getCell(temp, c))
        return columnList